import openpyxl
from openpyxl.utils import get_column_letter
file1=open('Chapter 13/text_to_spreadsheets/exampletext')
file2=open('Chapter 13/text_to_spreadsheets/example2text')
wb=openpyxl.Workbook()
sheet=wb.active
lst1=file1.readlines()
lst2=file2.readlines()
cumulative_list=(lst1,lst2)
total_lists=2

for i in range(1,total_lists+1):
    for j in range(1,len(cumulative_list[i-1])+1):
        r=get_column_letter(i)+str(j)
        sheet[r].value=lst1[j-1]

wb.save('converted_spreadsheet.xlsx')







