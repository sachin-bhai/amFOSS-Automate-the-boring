import openpyxl
from openpyxl.utils import get_column_letter


wb=openpyxl.load_workbook('Chapter 13/example.xlsx')

sheet=wb.active

r=sheet.max_row
c=sheet.max_column
text1=[]
text2=[]
texts=[text1,text2]
for i in range(1,c+1):
    for k in range(1,r+1):
        info=get_column_letter(i)+str(k)
        text=sheet[info].value
        texts[i-1].append(text)
print(text1)
print(text2)

file1=open('file.txt','w')
for i in text1:
    file1.write(str(i)+" ")
file1.close()

file2=open('file2.txt','w')
for i in text2:
    file2.write(str(i)+" ")
file2.close()



