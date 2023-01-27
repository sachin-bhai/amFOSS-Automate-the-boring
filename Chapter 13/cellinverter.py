
import openpyxl
from openpyxl.utils import get_column_letter
wb=openpyxl.load_workbook('Chapter 13/example.xlsx')
sheet=wb.active
row1=[]
row2=[]
rowmax=sheet.max_row
colmax=sheet.max_column
# print(rowmax,colmax)

for i in range(1,colmax+1):
    for k in range(1,rowmax+1):
        r1=get_column_letter(i)+str(k)
        c=sheet[r1]
        if i==1:
            row1.append(c.value)
        else:
            row2.append(c.value)

print(row1,row2)

wb.create_sheet(index=0,title='Sorted_sheet')
sheet=wb.active

for a in range(1,colmax+1):
    for b in range(1,rowmax+1):
        r2=get_column_letter(b)+str(a)
        if a==1:
            (sheet[r2]).value=row1[b-1]
        else:
            (sheet[r2]).value=row2[b-1]


wb.save('cell_inverter.xlsx')        







